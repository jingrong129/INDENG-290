from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from oil_forecast_project.config import RAW_DIR


CURRENT_AEO_URL = "https://www.eia.gov/outlooks/aeo/tables_ref.php"
ARCHIVE_AEO_URL = "https://www.eia.gov/outlooks/archive/aeo{yy}/tables_ref.php"
EIA_HEADERS = {"User-Agent": "Mozilla/5.0"}
EIA_RELEASE_DATES = {
    2013: "2013-04-15",
    2014: "2014-05-07",
    2015: "2015-04-14",
    2016: "2016-09-15",
    2017: "2017-01-05",
    2018: "2018-02-06",
    2019: "2019-01-24",
    2020: "2020-01-29",
    2021: "2021-02-03",
    2022: "2022-03-03",
    2023: "2023-03-16",
    2025: "2025-04-15",
    2026: "2026-04-08",
}


@dataclass(frozen=True)
class EiaVintageSpec:
    vintage_year: int
    page_url: str


def _page_url_for_vintage(vintage_year: int) -> str:
    if vintage_year == 2026:
        return CURRENT_AEO_URL
    return ARCHIVE_AEO_URL.format(yy=str(vintage_year)[-2:])


def list_eia_vintages(start_year: int = 2013, end_year: int = 2026) -> list[EiaVintageSpec]:
    return [
        EiaVintageSpec(vintage_year=year, page_url=_page_url_for_vintage(year))
        for year in range(start_year, end_year + 1)
        if year in EIA_RELEASE_DATES
    ]


def _get(url: str) -> requests.Response:
    response = requests.get(url, headers=EIA_HEADERS, timeout=60)
    response.raise_for_status()
    return response


def _extract_release_date(html: str) -> pd.Timestamp:
    candidates = []
    for pattern in [
        r"Release Dates?:.*?([A-Za-z]+\s+\d{1,2})(?:\s*-\s*[A-Za-z]+\s+\d{1,2})?,\s*(\d{4})",
        r"Release Dates?:.*?([A-Za-z]+\s+\d{1,2},\s+\d{4})",
        r"Release date:.*?([A-Za-z]+\s+\d{1,2},\s+\d{4})",
    ]:
        match = re.search(pattern, html, flags=re.IGNORECASE | re.DOTALL)
        if not match:
            continue
        if len(match.groups()) == 2:
            candidates.append(f"{match.group(1)}, {match.group(2)}")
        else:
            candidates.append(match.group(1))
    if not candidates:
        soup = BeautifulSoup(html, "html.parser")
        for container in soup.select(".release-dates"):
            date_spans = container.select("span.date")
            for date_span in date_spans:
                text = " ".join(date_span.get_text(" ", strip=True).split())
                if re.match(r"[A-Za-z]+\s+\d{1,2},\s+\d{4}", text):
                    candidates.append(text)
                    break
            if candidates:
                break
    if not candidates:
        raise ValueError("Unable to parse EIA release date.")
    return pd.to_datetime(candidates[0])


def _extract_release_date_for_page(page_url: str, html: str) -> pd.Timestamp:
    try:
        return _extract_release_date(html)
    except ValueError:
        base_url = page_url.rsplit("/", 1)[0] + "/"
        parent_html = _get(base_url).text
        return _extract_release_date(parent_html)


def _default_table12_xlsx_url(vintage_year: int) -> str:
    if vintage_year == 2026:
        return "https://www.eia.gov/outlooks/aeo/excel/aeotab12.xlsx"
    base = f"https://www.eia.gov/outlooks/archive/aeo{str(vintage_year)[-2:]}/excel/"
    filename = "aeotab12.xlsx" if vintage_year >= 2022 else "aeotab_12.xlsx"
    return urljoin(base, filename)


def _extract_table12_xlsx_url(vintage_year: int, page_url: str, html: str) -> str:
    anchor_pattern = re.compile(
        r"Table 12\.\s*Petroleum(?: and Other Liquids)? Prices.*?XLSX",
        re.IGNORECASE | re.DOTALL,
    )
    if not anchor_pattern.search(html):
        return _default_table12_xlsx_url(vintage_year)
    anchors = re.findall(r'<a href="([^"]+)">([^<]*)</a>', html, flags=re.IGNORECASE)
    for idx, (_, text) in enumerate(anchors):
        text_clean = " ".join(text.split())
        if text_clean.startswith("Table 12."):
            for href, next_text in anchors[idx + 1 : idx + 4]:
                if next_text.strip().upper() in {"XLSX", "XLS"}:
                    return urljoin(page_url, href)
    return _default_table12_xlsx_url(vintage_year)


def _detect_year_row(sheet: pd.DataFrame) -> int:
    for idx in range(min(20, len(sheet))):
        values = [value for value in sheet.iloc[idx].tolist() if pd.notna(value)]
        year_hits = 0
        for value in values:
            if isinstance(value, (int, float)) and 1900 <= int(value) <= 2100:
                year_hits += 1
            elif hasattr(value, "year"):
                year_hits += 1
        if year_hits >= 6:
            return idx
    raise ValueError("Unable to find year header row in EIA table.")


def _extract_dollar_year(sheet: pd.DataFrame) -> int:
    for idx in range(min(20, len(sheet))):
        row_text = " ".join(str(value) for value in sheet.iloc[idx].tolist() if pd.notna(value))
        match = re.search(r"\((\d{4}) dollars per barrel\)", row_text, flags=re.IGNORECASE)
        if match:
            return int(match.group(1))
    raise ValueError("Unable to parse EIA dollar year.")


def _extract_brent_row(sheet: pd.DataFrame) -> pd.Series:
    mask = sheet.astype(str).apply(lambda col: col.str.contains("Brent Spot", case=False, na=False)).any(axis=1)
    if not mask.any():
        raise ValueError("Unable to locate Brent row in EIA table.")
    return sheet.loc[mask].iloc[0]


def _normalize_year(value: object) -> int:
    if hasattr(value, "year"):
        return int(value.year)
    return int(value)


def fetch_eia_brent_vintages(start_year: int = 2013, end_year: int = 2026) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for spec in list_eia_vintages(start_year=start_year, end_year=end_year):
        release_date = pd.Timestamp(EIA_RELEASE_DATES[spec.vintage_year])
        xlsx_url = _default_table12_xlsx_url(spec.vintage_year)

        raw_path = RAW_DIR / "eia" / f"aeo_{spec.vintage_year}_table12.xlsx"
        raw_path.parent.mkdir(parents=True, exist_ok=True)
        if raw_path.exists():
            xlsx_bytes = raw_path.read_bytes()
        else:
            xlsx_bytes = _get(xlsx_url).content
            raw_path.write_bytes(xlsx_bytes)

        workbook = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
        sheet_name = workbook.sheetnames[0]
        sheet = pd.read_excel(BytesIO(xlsx_bytes), sheet_name=sheet_name, header=None)
        year_row = _detect_year_row(sheet)
        brent_row = _extract_brent_row(sheet)
        dollar_year = _extract_dollar_year(sheet)

        for col_idx, year_value in enumerate(sheet.iloc[year_row].tolist()):
            if pd.isna(year_value):
                continue
            try:
                target_year = _normalize_year(year_value)
            except (TypeError, ValueError):
                continue
            if target_year < 1900 or target_year > 2100:
                continue
            forecast_value = pd.to_numeric(brent_row.iloc[col_idx], errors="coerce")
            if pd.isna(forecast_value):
                continue
            records.append(
                {
                    "provider": "EIA",
                    "vintage_year": spec.vintage_year,
                    "release_date": release_date.normalize(),
                    "table12_base_year": dollar_year,
                    "target_year": target_year,
                    "forecast_real_base": float(forecast_value),
                    "source_url": xlsx_url,
                }
            )
    return pd.DataFrame.from_records(records).sort_values(["vintage_year", "target_year"]).reset_index(drop=True)
